from openpyxl.utils import get_column_letter
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table
import logging

from excelStyles import apply_excel_colors
from tableStyling import create_table_data, apply_table_styles
from openpyxl.styles import Alignment, Border, Side

logging.basicConfig(level=logging.DEBUG)


def save_pdf(data, file_path, page_size, dynamic_cols, config):
    """Generate a PDF with the given data and specifications."""
    if data.empty:
        logging.warning("No data to display in the PDF.")
        return

    try:
        doc = SimpleDocTemplate(file_path, pagesize=page_size)
        table_data = create_table_data(data, dynamic_cols)
        table = Table(table_data)

        apply_table_styles(table, table_data, dynamic_cols, config)
        doc.build([table])
        logging.info("PDF generation completed successfully.")

    except Exception as e:
        logging.error(f"Error building PDF: {e}")


def apply_styles_excel(df, worksheet, config):
    """Apply styles to Excel worksheet."""
    set_row_heights(worksheet)
    set_column_widths(df, worksheet)
    apply_cell_alignment(df, worksheet, config)
    apply_cell_borders(worksheet)
    apply_excel_colors(worksheet, config)


def save_excel(data, file_path, dynamic_cols, config):
    """Create an Excel report from the given data."""
    if data.empty:
        logging.warning("No data to display in the Excel file.")
        return

    try:
        table_data = create_table_data(data, dynamic_cols)
        df = pd.DataFrame(table_data[1:], columns=table_data[0])

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Report', index=False)
            worksheet = writer.sheets['Report']

            apply_styles_excel(df, worksheet, config)

        logging.info("Excel generation completed successfully.")

    except Exception as e:
        logging.error(f"Error building Excel: {e}")


def set_row_heights(worksheet):
    normal_row_height = 25
    header_row_height = normal_row_height + 10

    worksheet.row_dimensions[1].height = header_row_height

    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = normal_row_height


def set_column_widths(df, worksheet):
    for col_idx, col in enumerate(df.columns, 1):
        max_length = max((len(str(x)) for x in df[col]), default=10)
        adjusted_width = max_length + 5
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def apply_cell_borders(worksheet):
    """Apply borders to all cells in the worksheet."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def apply_cell_alignment(df, worksheet, config):
    for col_idx, col in enumerate(df.columns, 1):
        for row in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            horizontal_alignment = config['alignment']['header'].lower() if row == 1 else (config['alignment']['global']
                                                                                           .lower())
            cell.alignment = Alignment(horizontal=horizontal_alignment, vertical='center')


def dynamic_columns_for_pdf(group_cols, agg_col):
    """Determine dynamic columns for the PDF based on grouping and aggregate columns."""
    return group_cols + [agg_col]
