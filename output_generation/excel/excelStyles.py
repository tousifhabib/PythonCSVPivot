import logging
from typing import Dict, Any, List, Tuple, Set
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import units
from reportlab.lib.pagesizes import letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def calculate_dynamic_page_size(table_width, table_height, config, paginate, num_rows):
    default_cell_height = 0.25 * units.inch
    header_footer_height = 0.55 * units.inch

    margins = config.get("margins", {"top": 1, "bottom": 1, "left": 1, "right": 1})

    estimated_content_height = num_rows * default_cell_height + header_footer_height

    total_width = table_width + (margins["left"] + margins["right"]) * units.inch
    total_height = estimated_content_height + (margins["top"] + margins["bottom"]) * units.inch

    min_page_size = config.get("min_page_size", letter)
    max_page_size = config.get("max_page_size", (letter[0] * 2, letter[1] * 2))

    if paginate:
        return (
            max(min(total_width, max_page_size[0]), min_page_size[0]),
            max(min(total_height, max_page_size[1]), min_page_size[1])
        )
    else:
        new_width = max(min(total_width, max_page_size[0]), min_page_size[0])
        new_height = max(total_height, min_page_size[1])
        return new_width, new_height


def create_fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def create_font(color: str, bold: bool = False) -> Font:
    return Font(color=color, bold=bold)


def style_cell(cell: Cell, fill: PatternFill, font: Font) -> None:
    cell.fill = fill
    cell.font = font


def get_style_mappings(config: Dict[str, Any]) -> Dict[str, Tuple[PatternFill, Font]]:
    return {key: (create_fill(value['background']), create_font(value['text'], key == 'header'))
            for key, value in config['colors'].items()}


def determine_row_style(row_values: List[Any], style_mappings: Dict[str, Tuple[PatternFill, Font]]) -> Tuple[
    Tuple[PatternFill, Font], int]:
    if "Grand Total" in row_values:
        return style_mappings['grand_total'], 1
    if any(keyword in row_values for keyword in ["Grand Total"]):
        return style_mappings['grand_total'], 1
    level = next((i for i, value in enumerate(row_values[:-2], 1) if value), 0)
    if level:
        return style_mappings.get(f'subtotal_{level}', style_mappings['default']), level
    return style_mappings['default'], 1


def is_special_row(row_values: List[Any]) -> bool:
    if any(keyword in row_values for keyword in ["Grand Total"]):
        return True
    return any(cell and (i == 1 or not row_values[i - 1]) for i, cell in enumerate(row_values[1:], start=1))


def apply_row_styles(worksheet: Worksheet, config: Dict[str, Any]) -> None:
    style_mappings = get_style_mappings(config)
    for row_index, row in enumerate(worksheet.iter_rows(), start=1):
        row_values = [cell.value for cell in row]
        if row_index == 1:
            fill, font = style_mappings['header']
            start_col = 1
        else:
            (fill, font), start_col = determine_row_style(row_values, style_mappings)
        for cell in row[start_col - 1:]:
            style_cell(cell, fill, font)


def apply_excel_colors(worksheet: Worksheet, config: Dict[str, Any]) -> None:
    apply_row_styles(worksheet, config)


def merge_empty_cells_new(worksheet, df: pd.DataFrame) -> None:
    subtotal_rows, subtotal_levels = identify_subtotal_rows_new(df)
    merge_cells_new(worksheet, df, subtotal_rows, subtotal_levels)


def identify_subtotal_rows_new(df: pd.DataFrame) -> Tuple[Set[int], Dict[int, int]]:
    subtotal_rows = set()
    subtotal_levels = {}
    for col_idx, col in enumerate(df.columns[:-2], start=1):
        for row_idx, value in enumerate(df[col], start=2):
            if not pd.isna(value) and value != '':
                subtotal_rows.add(row_idx)
                subtotal_levels[row_idx] = col_idx

    print(subtotal_rows)
    print(subtotal_levels)
    return subtotal_rows, subtotal_levels


def merge_cells_new(worksheet, df: pd.DataFrame, subtotal_rows: Set[int], subtotal_levels: Dict[int, int]) -> None:
    num_columns = len(df.columns)

    for col_idx in range(num_columns - 2, 0, -1):
        print(f"\n-- Processing Column: {col_idx}")
        start_row = None
        for idx, value in enumerate(df.iloc[:, col_idx - 1], start=2):
            print(f"\n- Processing row: {idx}")
            is_subtotal_row = idx in subtotal_rows and subtotal_levels[idx] <= col_idx
            is_empty_row = pd.isna(value) or value == ''

            if idx in subtotal_rows:
                correct_level_to_merge = subtotal_levels[idx] > col_idx
            else:
                correct_level_to_merge = True

            if start_row is not None and not(is_empty_row and correct_level_to_merge):
                print(f"Merge Column: {col_idx}, Rows: {start_row} to {idx - 1}")
                merge_cells(worksheet, start_row, col_idx, idx - 1)
                start_row = None

            elif is_empty_row and correct_level_to_merge and not is_subtotal_row:
                if start_row is None:
                    start_row = idx

        if start_row is not None:
            print(f"Merge Column: {col_idx}, Rows: {start_row} to {len(df) + 1}")
            merge_cells(worksheet, start_row, col_idx, len(df) + 1)


def handle_value_column_separately(worksheet, df, subtotal_rows, num_columns):
    start_row = None
    for row_idx, value in enumerate(df.iloc[:, -1], start=2):
        if pd.isna(value) or value == '':
            if start_row is None:
                start_row = row_idx
        else:
            if start_row is not None and row_idx - 1 != start_row:
                if row_idx - 1 not in subtotal_rows:
                    merge_cells(worksheet, start_row, num_columns, row_idx - 1)
            start_row = None
    if start_row is not None and start_row <= len(df):
        merge_cells(worksheet, start_row, num_columns, len(df) + 1)


def merge_cells(worksheet, start_row: int, col_idx: int, end_row: int) -> None:
    if end_row > start_row:
        worksheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
