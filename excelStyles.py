import logging
from typing import Dict, Any, List, Tuple, Set

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame


logging.basicConfig(level=logging.DEBUG, format='%(message)s')

def create_fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def create_font(color: str, bold: bool = False) -> Font:
    return Font(color=color, bold=bold)


def style_cell(cell: Cell, fill: PatternFill, font: Font) -> None:
    cell.fill = fill
    cell.font = font


def apply_row_styles(worksheet: Worksheet, config: Dict[str, Any], dynamic_cols: List[str]) -> None:
    style_mappings = {
        'header': (
            create_fill(config['colors']['header']['background']),
            create_font(config['colors']['header']['text'], True)),
        'default': (
            create_fill(config['colors']['default']['background']), create_font(config['colors']['default']['text'])),
        'subtotal_1': (
            create_fill(config['colors']['subtotal_1']['background']),
            create_font(config['colors']['subtotal_1']['text'])),
        'subtotal_2': (
            create_fill(config['colors']['subtotal_2']['background']),
            create_font(config['colors']['subtotal_2']['text'])),
        'grand_total': (create_fill(config['colors']['grand_total']['background']),
                        create_font(config['colors']['grand_total']['text']))
    }

    def determine_style(row_values: List[Any], idx: int, dynamic_cols: List[str]) -> Tuple[PatternFill, Font]:
        logging.debug(f"Processing row {idx}: {row_values}")

        if "Grand Total" in row_values:
            logging.debug(f"Row {idx} is a Grand Total row")
            return style_mappings['grand_total']
        elif is_special_row_excel(row_values, dynamic_cols):
            level = 0
            for i, value in enumerate(row_values[:-2], 1):
                if value:
                    level = i
                    break
            if level > 0:
                logging.debug(f"Row {idx} is a subtotal row of level {level}")
                return style_mappings.get(f'subtotal_{level}', style_mappings['subtotal_1'])
            else:
                logging.debug(f"Row {idx} is a normal row")
                return style_mappings['default']
        else:
            logging.debug(f"Row {idx} is a normal row")
            return style_mappings['default']

    for idx, row in enumerate(worksheet.iter_rows(), start=1):
        row_values = [cell.value for cell in row]
        print(row_values)
        fill, font = determine_style(row_values, idx, dynamic_cols) if idx > 1 else style_mappings['header']
        for cell in row:
            style_cell(cell, fill, font)


def is_special_row_excel(row_values: List[Any], dynamic_cols: List[str]) -> bool:
    if "Subtotal" in row_values or "Grand Total" in row_values:
        return True

    for i, cell in enumerate(row_values[1:], start=1):  # start from the second cell
        if cell:  # if the cell is not empty, check the previous one
            if i == 1 or (i > 1 and not row_values[i-1]):  # if the first cell is empty or previous cell is empty
                return True
    return False


def apply_excel_colors(worksheet: Worksheet, config: Dict[str, Any], df: DataFrame, dynamic_cols: List[str]) -> None:
    apply_row_styles(worksheet, config, dynamic_cols)


def merge_empty_cells(worksheet, df: pd.DataFrame) -> None:
    subtotal_rows, subtotal_levels = identify_subtotal_rows(df)
    merge_first_column(worksheet, subtotal_rows, subtotal_levels)
    merge_remaining_columns(worksheet, df, subtotal_rows, subtotal_levels)


def identify_subtotal_rows(df: pd.DataFrame) -> Tuple[Set[int], Dict[int, int]]:
    subtotal_rows = set()
    subtotal_levels = {}
    for col_idx, col in enumerate(df.columns[:-2], start=1):
        blank_seen = False
        level = 0
        for row_idx, value in enumerate(df[col], start=2):
            if pd.isna(value) or value == '':
                blank_seen = True
            else:
                if blank_seen or "Grand Total" in str(value):
                    subtotal_rows.add(row_idx)
                    if col_idx == 1:
                        subtotal_levels[row_idx] = level
                    blank_seen = False
                    level += 1
    return subtotal_rows, subtotal_levels


def merge_first_column(worksheet, subtotal_rows: Set[int], subtotal_levels: Dict[int, int]) -> None:
    row_idx_list = sorted(subtotal_rows)
    start_row = 3
    prev_level = -1
    for row_idx in row_idx_list:
        level = subtotal_levels.get(row_idx, -1)
        if level > prev_level and row_idx > start_row:
            merge_cells(worksheet, start_row, 1, row_idx - 1)
            start_row = row_idx + 1
        prev_level = level
    if start_row <= row_idx_list[-1]:
        merge_cells(worksheet, start_row, 1, row_idx_list[-1])


def merge_remaining_columns(worksheet, df: pd.DataFrame, subtotal_rows: Set[int],
                            subtotal_levels: Dict[int, int]) -> None:
    for col_idx, col in enumerate(df.columns, start=1):
        start_row = None
        for row_idx, value in enumerate(df[col], start=2):
            is_subtotal_row = row_idx in subtotal_rows
            if is_subtotal_row or not pd.isna(value) and value != '':
                if start_row is not None and row_idx - 1 != start_row and col_idx > subtotal_levels.get(row_idx, 0) + 1:
                    merge_cells(worksheet, start_row, col_idx, row_idx - 1)
                start_row = None
            elif start_row is None:
                start_row = row_idx
        if start_row is not None and start_row <= len(df):
            merge_cells(worksheet, start_row, col_idx, len(df) + 1)


def merge_cells(worksheet, start_row: int, col_idx: int, end_row: int) -> None:
    if end_row > start_row:
        worksheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
